"""backfill_traders.py - One-time historical import of trader-bid history.

Walks the Trader Bids workbook (one tab per calendar year) and inserts the
WINNING bid for each weekly box into guild_traders. The winning bid is the
row that's been formatted bold by the GM (with a colored fill) - openpyxl's
font/fill metadata makes that detectable from the xlsx.

Layout differences across years:
  * 2024-2026: 5 columns per box: [location, rank, trader, amount, sep].
                Header row: [start_date, None, 'to', end_date, None].
  * 2022-2023: 4 columns per box: [location, trader, amount, sep].
                Header row: single string 'MM/DD - MM/DD' in col 1.
  * Pre-2022: too inconsistent; user requested skipping these.

Each year tab has 4 boxes per row, with header rows separating row-groups.
A box may have NO winning bid (if no bid won that week); we simply skip those.

Idempotent: upserts on week_id, so re-running is safe.

Usage:
    python backfill_traders.py --db guildstats.db \\
        --xlsx "Trader Bids.xlsx" \\
        --years 2022,2023,2024,2025,2026
"""
from __future__ import annotations
import argparse
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from guildstats import (
    open_db, ingest_run, upsert_week, upsert_trader_bid,
    TraderBid, trade_week_ending,
)


# Box layouts per year. Each entry is the column offset of the FIRST column
# of each of the 4 boxes (1-indexed for openpyxl).
LAYOUTS = {
    "modern": dict(box_starts=[1, 6, 11, 16], box_width=5,
                   loc_off=0, rank_off=1, trader_off=2, amount_off=3,
                   header_start_off=0, header_end_off=3),
    "classic": dict(box_starts=[1, 5, 9, 13], box_width=4,
                    loc_off=0, trader_off=1, amount_off=2,
                    header_combined_off=0),
}


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _is_bold(ws, row: int, col_start: int, box_width: int) -> bool:
    """True if any cell in this box's span on this row has bold font."""
    for c in range(col_start, col_start + box_width):
        cell = ws.cell(row=row, column=c)
        if cell.font and cell.font.bold:
            return True
    return False


def _classify_layout(year: int) -> dict:
    return LAYOUTS["classic"] if year <= 2023 else LAYOUTS["modern"]


# Date parsing for the legacy "MM/DD - MM/DD" string headers
_RANGE_RE = re.compile(
    r"\s*(\d{1,2})[/-](\d{1,2})\s*[-]\s*(\d{1,2})[/-](\d{1,2})\s*"
)


def _parse_classic_header(text: str, year: int):
    """Parse '01/03 - 01/10' -> end-date-of-week as a date.
    Returns datetime or None."""
    if text is None:
        return None
    m = _RANGE_RE.match(str(text))
    if not m:
        return None
    end_mm, end_dd = int(m.group(3)), int(m.group(4))
    try:
        return datetime(year, end_mm, end_dd, 19, 0, 0, tzinfo=timezone.utc)
    except ValueError:
        return None


def _modern_header_dt(ws, row: int, layout: dict, box_idx: int):
    """Get the end_date from a modern-layout header row."""
    col = layout["box_starts"][box_idx] + layout["header_end_off"]
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(hour=19, minute=0, second=0, microsecond=0,
                           tzinfo=timezone.utc) if val.tzinfo is None else val
    # Could be a formula result rendered as a string; try parsing
    try:
        d = datetime.strptime(str(val).split()[0], "%Y-%m-%d")
        return d.replace(hour=19, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        return None


def _is_modern_header_row(ws, row: int, layout: dict, box_idx: int) -> bool:
    """True if this row, in this box, has the 'start_date / to / end_date' header pattern."""
    col_start = layout["box_starts"][box_idx]
    val_start = ws.cell(row=row, column=col_start).value
    val_to = ws.cell(row=row, column=col_start + 2).value
    if val_to is None:
        return False
    return str(val_to).strip().lower() == "to" and val_start is not None


def _is_classic_header_row(ws, row: int, layout: dict, box_idx: int) -> bool:
    col_start = layout["box_starts"][box_idx]
    val = ws.cell(row=row, column=col_start).value
    if val is None or not isinstance(val, str):
        return False
    return _RANGE_RE.match(val) is not None


def backfill_year(conn, ws, year: int, counts: dict, errors: list) -> None:
    layout = _classify_layout(year)
    is_classic = (year <= 2023)
    is_header = _is_classic_header_row if is_classic else _is_modern_header_row

    print(f"\n=== {year} (layout={'classic' if is_classic else 'modern'}) ===")

    # For each of the 4 box positions, walk the rows independently.
    for box_idx in range(4):
        col_start = layout["box_starts"][box_idx]
        # Find header rows in this box
        header_rows = []
        last_row = ws.max_row
        for r in range(1, last_row + 1):
            if is_header(ws, r, layout, box_idx):
                header_rows.append(r)
        # For each header, scan rows below it (until the next header or last_row)
        # for the bold winning row. Skip the totals/sum row.
        for i, hdr_row in enumerate(header_rows):
            scan_end = header_rows[i + 1] if i + 1 < len(header_rows) else last_row + 1
            # Get end_date for this box
            if is_classic:
                hdr_text = ws.cell(row=hdr_row,
                                   column=col_start + layout["header_combined_off"]).value
                end_dt = _parse_classic_header(hdr_text, year)
            else:
                end_dt = _modern_header_dt(ws, hdr_row, layout, box_idx)

            if end_dt is None:
                continue

            # Find the bold row in the scan range
            winner_row = None
            for r in range(hdr_row + 1, scan_end):
                if not _is_bold(ws, r, col_start, layout["box_width"]):
                    continue
                # Confirm it's a data row, not a totals-only row
                loc = ws.cell(row=r, column=col_start + layout["loc_off"]).value
                if is_classic:
                    trader = ws.cell(row=r, column=col_start + layout["trader_off"]).value
                else:
                    trader = ws.cell(row=r, column=col_start + layout["trader_off"]).value
                amount = _to_int(ws.cell(row=r,
                                         column=col_start + layout["amount_off"]).value)
                if loc and trader and amount and amount > 0:
                    winner_row = r
                    break
                # Otherwise it's likely a sum/total row that's also bold; keep scanning

            if winner_row is None:
                # No winning bid this week (entirely valid; user noted this happens)
                continue

            location = str(ws.cell(row=winner_row,
                                   column=col_start + layout["loc_off"]).value).strip()
            trader_name = str(ws.cell(row=winner_row,
                                      column=col_start + layout["trader_off"]).value).strip()
            amount = _to_int(ws.cell(row=winner_row,
                                     column=col_start + layout["amount_off"]).value)

            # Snap end_dt to the canonical Tuesday 19:00 UTC for the trade week
            week_end = trade_week_ending(end_dt)
            try:
                week_id = upsert_week(conn, week_end)
                upsert_trader_bid(conn, TraderBid(
                    week_id=week_id,
                    trader_name=trader_name,
                    location=location,
                    bid_amount=amount,
                    notes=None,
                ))
                counts["rows_inserted"] += 1
            except Exception as exc:
                errors.append(f"  {year} box{box_idx} R{winner_row}: {exc}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", required=True)
    ap.add_argument("--xlsx", required=True, help="Trader Bids.xlsx")
    ap.add_argument("--years", default="2022,2023,2024,2025,2026",
                    help="Comma-separated list of year tabs to import")
    args = ap.parse_args()

    target_years = [int(y.strip()) for y in args.years.split(",")]

    conn = open_db(args.db)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    errors: list = []
    with ingest_run(conn, "backfill_traders",
                    workbook_filename=args.xlsx) as counts:
        conn.execute("BEGIN")
        try:
            for year in target_years:
                tab = str(year)
                if tab not in wb.sheetnames:
                    print(f"SKIP: tab {tab!r} not found in workbook")
                    continue
                backfill_year(conn, wb[tab], year, counts, errors)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise

    if errors:
        print(f"\n=== {len(errors)} warnings ===")
        for e in errors[:25]:
            print(e)
        if len(errors) > 25:
            print(f"  ... and {len(errors) - 25} more")
    print(f"\nDone. {counts['rows_inserted']} trader bids upserted.")


if __name__ == "__main__":
    main()
