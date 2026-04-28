"""import_winners.py - Mirror prizes/winners from the raffle workbooks into
the database after the legacy Apps Script has finished its Friday rollover.

The Apps Script renames the 'Current' tab to MMDDYY and creates a fresh empty
'Current'. So we read the LATEST dated tab (the one that just got renamed)
rather than 'Current' itself.

Also creates/updates raffle_entries from that tab so the database reflects
the final entry list after donations were promoted into the spreadsheet.

Usage:
  python import_winners.py --db guildstats.db \\
      --standard "AKTT Standard Raffle.xlsx" \\
      --highroller "AKTT High-Roller Raffle.xlsx"

  # Or import a specific tab by name:
  python import_winners.py --db guildstats.db \\
      --standard "AKTT Standard Raffle.xlsx" --tab 042426
"""
from __future__ import annotations
import argparse
import re
import sys
from datetime import datetime, timezone

import openpyxl

from guildstats import (
    open_db, ingest_run,
    upsert_user, upsert_raffle, insert_raffle_entry,
    upsert_prize, upsert_winner, recompute_raffle_totals,
    RaffleEntry, Prize,
)

# Reuse the parsing pieces from backfill_raffle so behavior matches.
from backfill_raffle import (
    DATE_TAB_RE, _parse_dt,
    _parse_entry_std, _parse_entry_hr,
    _parse_main_prizes_rows, _parse_mini_prizes_std_rows,
    STD_PRIZE_COLS, STD_PRIZE_ROW_RANGE,
    HR_PRIZE_COLS, HR_PRIZE_ROW_RANGE,
)


def _latest_dated_tab(workbook) -> str | None:
    """Return the most-recent MMDDYY-named tab by drawing_date, or None."""
    candidates = []
    for name in workbook.sheetnames:
        if not DATE_TAB_RE.fullmatch(name):
            continue
        # Convert MMDDYY -> sortable YYYYMMDD assuming 20YY
        mm, dd, yy = name[:2], name[2:4], name[4:6]
        sortkey = f"20{yy}{mm}{dd}"
        candidates.append((sortkey, name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _import_one_tab(conn, ws, raffle_type: str, tab_name: str) -> dict:
    """Re-import entries + prizes + winners for a single tab. Returns counts."""
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return {"entries": 0, "prizes": 0, "winners": 0}

    row2 = all_rows[1] if len(all_rows) >= 2 else None
    drawing_dt = _parse_dt(row2[10] if row2 and len(row2) > 10 else None)
    if drawing_dt is None:
        raise SystemExit(f"Tab {tab_name} has no drawing date in K2")

    raffle_id = upsert_raffle(conn, raffle_type=raffle_type,
                              drawing_date=drawing_dt.date(),
                              status="drawn")

    parse_entry = _parse_entry_std if raffle_type == "standard" else _parse_entry_hr

    # Re-import entries (idempotent: insert_raffle_entry dedupes by source_transaction_id;
    # other entries may insert duplicates, so we scrub manually-source entries first).
    # For safety we wipe non-bank entries on this raffle, then re-insert.
    conn.execute("DELETE FROM raffle_entries WHERE raffle_id=? "
                 "AND source IN ('mail_donation', 'event')", (raffle_id,))

    entries_added = 0
    for r_idx in range(3, min(203, len(all_rows))):
        row = all_rows[r_idx]
        if not row:
            continue
        e = parse_entry(row, raffle_id, conn)
        if e is None:
            continue
        # is_backfilled remains False here since this is post-drawing live import,
        # not from the historical archive.
        e.is_backfilled = False
        insert_raffle_entry(conn, e)
        entries_added += 1

    if raffle_type == "standard":
        main_p, main_w = _parse_main_prizes_rows(all_rows, raffle_id, conn,
                                                 STD_PRIZE_COLS, STD_PRIZE_ROW_RANGE)
        mini_p, mini_w = _parse_mini_prizes_std_rows(all_rows, raffle_id, conn)
        prize_count = main_p + mini_p
        winner_count = main_w + mini_w
    else:
        main_p, main_w = _parse_main_prizes_rows(all_rows, raffle_id, conn,
                                                 HR_PRIZE_COLS, HR_PRIZE_ROW_RANGE)
        prize_count = main_p
        winner_count = main_w

    recompute_raffle_totals(conn, raffle_id)
    print(f"  {raffle_type:12} tab {tab_name} (drawing {drawing_dt.date()}): "
          f"{entries_added} entries, {prize_count} prizes, {winner_count} winners")
    return {"entries": entries_added, "prizes": prize_count, "winners": winner_count}


def import_workbook(conn, path: str, raffle_type: str, tab: str | None) -> dict:
    print(f"\n=== {raffle_type} from {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if tab is None:
        tab = _latest_dated_tab(wb)
        if tab is None:
            raise SystemExit(f"No dated tabs found in {path}")
        print(f"  using latest dated tab: {tab}")
    else:
        if tab not in wb.sheetnames:
            raise SystemExit(f"Tab '{tab}' not found in {path}")
    return _import_one_tab(conn, wb[tab], raffle_type, tab)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", required=True)
    ap.add_argument("--standard", help="AKTT Standard Raffle.xlsx")
    ap.add_argument("--highroller", help="AKTT High-Roller Raffle.xlsx")
    ap.add_argument("--tab", default=None,
                    help="Specific tab name (MMDDYY); default = latest dated tab")
    args = ap.parse_args()
    if not args.standard and not args.highroller:
        ap.error("at least one of --standard / --highroller is required")

    conn = open_db(args.db)
    notes = []
    with ingest_run(conn, "import_winners",
                    workbook_filename=args.standard or args.highroller) as counts:
        conn.execute("BEGIN")
        try:
            if args.standard:
                r = import_workbook(conn, args.standard, "standard", args.tab)
                counts["rows_inserted"] += r["entries"] + r["prizes"] + r["winners"]
                notes.append(f"std: {r}")
            if args.highroller:
                r = import_workbook(conn, args.highroller, "high_roller", args.tab)
                counts["rows_inserted"] += r["entries"] + r["prizes"] + r["winners"]
                notes.append(f"hr: {r}")
            counts["notes"] = "; ".join(notes)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise
    print("\nDone.")


if __name__ == "__main__":
    main()
