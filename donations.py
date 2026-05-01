"""donations.py - CLI for entering and managing mail-in auction donations.

Subcommands:
  add        Record a donation: user, value, optional description
  list       List donations (defaults to current trade week, unpromoted only)
  promote    Convert unpromoted donations from a trade week into raffle_entries
             on the next standard raffle (mirrors the Tuesday-rollover Apps Script)

Examples:
  python donations.py --db guildstats.db add @user 60000 "gold mats"
  python donations.py --db guildstats.db list
  python donations.py --db guildstats.db list --week 2026-04-21 --all
  python donations.py --db guildstats.db promote --to-raffle current
  python donations.py --db guildstats.db promote --week 2026-04-21 --to-raffle 47
"""
from __future__ import annotations
import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from guildstats import (
    open_db, upsert_user, upsert_week, trade_week_ending,
    add_manual_donation, promote_donations_to_raffle,
    find_open_raffle, raffle_drawing_date_for, upsert_raffle,
    DONATION_MIN_VALUE,
)


def _resolve_week_id(conn, week_arg: str | None) -> int:
    """Resolve --week. None means 'current trade week'. A YYYY-MM-DD value
    means 'the trade week ending that date'. Always upserts so it's safe."""
    if week_arg is None:
        end = trade_week_ending(datetime.now(timezone.utc))
    else:
        end_date = datetime.strptime(week_arg, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        # Snap to canonical Tuesday 19:00 UTC
        end = end_date.replace(hour=19, minute=0, second=0, microsecond=0)
    return upsert_week(conn, end)


def _resolve_raffle_id(conn, raffle_arg: str) -> int:
    """Resolve --to-raffle. 'current' means the next open standard raffle
    (creating one if necessary). A numeric id means that exact raffle."""
    if raffle_arg.lower() == "current":
        rid = find_open_raffle(conn, raffle_type="standard")
        if rid is not None:
            return rid
        # No open raffle — create one for the upcoming Friday
        drawing = raffle_drawing_date_for(datetime.now(timezone.utc))
        return upsert_raffle(conn, raffle_type="standard",
                             drawing_date=drawing, status="open")
    return int(raffle_arg)


def cmd_add(conn, args):
    user_id = upsert_user(conn, args.user)
    week_id = _resolve_week_id(conn, args.week)
    received_at = (datetime.strptime(args.received_at, "%Y-%m-%d %H:%M:%S")
                   .replace(tzinfo=timezone.utc)) if args.received_at else None
    don_id = add_manual_donation(conn, user_id=user_id, week_id=week_id,
                                 value=args.value, description=args.description,
                                 received_at=received_at,
                                 recorded_by=args.recorded_by)
    print(f"Added donation #{don_id}: {args.user} {args.value:,} gold "
          f"({args.description or 'no description'})")


def cmd_list(conn, args):
    if args.all_weeks:
        where = "1=1"
        params: tuple = ()
    else:
        week_id = _resolve_week_id(conn, args.week)
        where = "md.week_id = ?"
        params = (week_id,)
    promo_clause = "" if args.show_promoted else "AND md.is_promoted = 0"
    rows = conn.execute(
        f"""
        SELECT md.id, u.account_name, w.ending_date, md.value, md.description,
               md.is_promoted, md.promoted_to_raffle_id, md.received_at, md.recorded_by
          FROM manual_donations md
          JOIN users u ON u.id = md.user_id
          JOIN weeks w ON w.id = md.week_id
         WHERE {where} {promo_clause}
         ORDER BY md.received_at DESC, md.id DESC
        """, params).fetchall()
    if not rows:
        print("(no donations matched)")
        return
    print(f"{'ID':>5}  {'USER':<22} {'WEEK ENDING':<11}  {'VALUE':>10}  PROMOTED  DESCRIPTION")
    for r in rows:
        promo = f"r#{r['promoted_to_raffle_id']}" if r["is_promoted"] else "no"
        desc = (r["description"] or "")[:40]
        print(f"{r['id']:>5}  {r['account_name']:<22} {r['ending_date']:<11}  "
              f"{r['value']:>10,}  {promo:<8}  {desc}")
    total = sum(r["value"] for r in rows)
    print(f"\n{len(rows)} rows, total {total:,} gold")


def cmd_promote(conn, args):
    week_id = _resolve_week_id(conn, args.week)
    raffle_id = _resolve_raffle_id(conn, args.to_raffle)
    raffle_row = conn.execute(
        "SELECT raffle_type, drawing_date FROM raffles WHERE id=?",
        (raffle_id,)).fetchone()
    if not raffle_row:
        print(f"ERROR: raffle id {raffle_id} not found", file=sys.stderr)
        sys.exit(2)
    print(f"Promoting unpromoted donations from week_id={week_id} "
          f"into {raffle_row['raffle_type']} raffle id={raffle_id} "
          f"(drawing {raffle_row['drawing_date']})")
    promoted = promote_donations_to_raffle(conn, week_id=week_id,
                                           raffle_id=raffle_id,
                                           min_value=args.min_value)
    print(f"Promoted {promoted} user(s).")



def cmd_import_from_sheet(conn, args):
    """Sync the latest dated tab of the Auction Donations workbook into
    manual_donations rows (source='sheet_import').

    Idempotent: re-running on the same sheet content is a no-op. If a row's
    content changes in the sheet, the corresponding manual_donations row is
    re-imported (we delete unpromoted sheet_import rows for the target week
    before re-inserting). Already-promoted rows are never touched.
    """
    import hashlib
    import re
    from datetime import datetime, timezone
    from pathlib import Path
    import openpyxl

    # Download the workbook from Drive (or use --xlsx for offline use)
    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.is_file():
            raise SystemExit(f"--xlsx file not found: {xlsx_path}")
    else:
        from drive_sync import export_sheet_as_xlsx, resolve_id
        spreadsheet_id = resolve_id("donations", args.spreadsheet_id)
        xlsx_path = Path(args.out or "/tmp/aktt-donations.xlsx")
        export_sheet_as_xlsx(spreadsheet_id, xlsx_path, key_path=args.key)
        print(f"Downloaded donations workbook -> {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # Pick which tab to import. Default = latest MMDDYY tab.
    DATE_TAB_RE = re.compile(r"^\d{6}$")
    if args.tab:
        tab = args.tab
        if tab not in wb.sheetnames:
            raise SystemExit(f"Tab '{tab}' not found in workbook")
    else:
        candidates = []
        for name in wb.sheetnames:
            if DATE_TAB_RE.fullmatch(name):
                mm, dd, yy = name[:2], name[2:4], name[4:6]
                candidates.append((f"20{yy}{mm}{dd}", name))
        if not candidates:
            raise SystemExit("No MMDDYY tabs found in workbook")
        candidates.sort(reverse=True)
        tab = candidates[0][1]
        print(f"Using latest dated tab: {tab}")

    ws = wb[tab]

    # Resolve the trade-week-ending date from the tab name (MMDDYY).
    mm, dd, yy = tab[:2], tab[2:4], tab[4:6]
    tab_date = datetime.strptime(f"20{yy}-{mm}-{dd}", "%Y-%m-%d").replace(tzinfo=timezone.utc)
    week_id = _resolve_week_id(conn, tab_date.date().isoformat())

    # Replace existing unpromoted sheet_import rows for this week
    deleted = conn.execute(
        "DELETE FROM manual_donations WHERE week_id = ? "
        "AND source = 'sheet_import' AND is_promoted = 0",
        (week_id,)).rowcount
    if deleted:
        print(f"Removed {deleted} stale sheet_import rows for this week")

    # Walk rows 4-40 (the data range used by the legacy Apps Script)
    inserted = skipped = 0
    received_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    for row in ws.iter_rows(min_row=4, max_row=40, values_only=True):
        if len(row) < 2:
            continue
        name = row[0]
        value = row[1]
        description = row[2] if len(row) > 2 else None
        if name is None or value is None:
            continue
        name = str(name).strip()
        if not name.startswith("@"):
            skipped += 1
            continue
        try:
            value_int = int(float(value))
        except (TypeError, ValueError):
            skipped += 1
            continue
        if value_int <= 0:
            skipped += 1
            continue
        desc_str = str(description).strip() if description else None
        # Hash the (name, value, description) triple for idempotency / debugging
        hash_input = f"{tab}|{name}|{value_int}|{desc_str or ''}"
        row_hash = hashlib.sha1(hash_input.encode("utf-8")).hexdigest()[:16]

        user_id = upsert_user(conn, name)
        conn.execute(
            "INSERT INTO manual_donations "
            "  (user_id, week_id, value, description, received_at, recorded_by, "
            "   source, sheet_row_hash) "
            "VALUES (?, ?, ?, ?, ?, ?, 'sheet_import', ?)",
            (user_id, week_id, value_int, desc_str, received_at,
             "sheet_sync", row_hash),
        )
        inserted += 1

    print(f"Imported {inserted} donation rows from {tab} (skipped {skipped})")

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", required=True)
    sub = ap.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("add", help="Record one mail-in donation")
    a.add_argument("user", help="Account name e.g. @jeffk42")
    a.add_argument("value", type=int, help="Estimated value in gold")
    a.add_argument("description", nargs="?", default=None,
                   help="Free-text description of items donated")
    a.add_argument("--week", default=None,
                   help="Trade-week ending date YYYY-MM-DD; default = current week")
    a.add_argument("--received-at", default=None,
                   help="When the items were received: 'YYYY-MM-DD HH:MM:SS' UTC; default = now")
    a.add_argument("--recorded-by", default=None,
                   help="Officer name for audit trail")
    a.set_defaults(func=cmd_add)

    l = sub.add_parser("list", help="List donations")
    l.add_argument("--week", default=None,
                   help="Trade-week ending date YYYY-MM-DD; default = current")
    l.add_argument("--all", dest="all_weeks", action="store_true",
                   help="Show donations from every week, ignoring --week")
    l.add_argument("--show-promoted", action="store_true",
                   help="Include donations that have already been promoted")
    l.set_defaults(func=cmd_list)

    p = sub.add_parser("promote", help="Promote a week's donations into a raffle")
    p.add_argument("--week", default=None,
                   help="Trade-week ending date YYYY-MM-DD; default = current")
    p.add_argument("--to-raffle", default="current",
                   help="Raffle id, or 'current' for the next open standard raffle")
    p.add_argument("--min-value", type=int, default=DONATION_MIN_VALUE,
                   help="Per-user minimum aggregated value to promote")
    p.set_defaults(func=cmd_promote)

    s = sub.add_parser("import-from-sheet",
                       help="Sync donations from the Google Sheet into manual_donations")
    s.add_argument("--tab", default=None,
                   help="MMDDYY tab name; default = latest dated tab")
    s.add_argument("--xlsx", default=None,
                   help="Path to a local xlsx (skips Drive download)")
    s.add_argument("--spreadsheet-id", default=None,
                   help="Override AKTT_DRIVE_DONATIONS_ID")
    s.add_argument("--key", default=None,
                   help="Override AKTT_DRIVE_KEY (service-account JSON path)")
    s.add_argument("--out", default=None,
                   help="Where to save the downloaded xlsx (default: /tmp/aktt-donations.xlsx)")
    s.set_defaults(func=cmd_import_from_sheet)

    args = ap.parse_args()
    conn = open_db(args.db)
    conn.execute("BEGIN")
    try:
        args.func(conn, args)
        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise


if __name__ == "__main__":
    main()
