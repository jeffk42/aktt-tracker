"""backfill_raffle.py - Historical backfill of raffles, entries, prizes and
winners from the two raffle workbooks.

Reads:
  * AKTT Standard Raffle.xlsx     -> raffle_type='standard'
  * AKTT High-Roller Raffle.xlsx  -> raffle_type='high_roller'

Each MMDDYY tab becomes one raffles row plus its entries, prizes, and
raffle_winners. Standard raffles also pick up the mini-prize block at
rows 59-68. Tabs that fail to parse are skipped with a warning.

Usage:
    python backfill_raffle.py --db guildstats.db \
        --standard "AKTT Standard Raffle.xlsx" \
        --highroller "AKTT High-Roller Raffle.xlsx" \
        --schema schema.sql
"""
from __future__ import annotations
import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from guildstats import (
    open_db, apply_schema, ingest_run,
    upsert_user, upsert_raffle, insert_raffle_entry,
    upsert_prize, upsert_winner, recompute_raffle_totals,
    RaffleEntry, Prize,
)

DATE_TAB_RE = re.compile(r"^\d{6}$")


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _norm_account(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s.startswith("@") else None


def _parse_dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def _parse_donation_value_from_date_column(v):
    """For DONATION rows the integer value lives in the date column. Excel
    often formats large integers as dates - recover the int either way."""
    if v is None:
        return None
    if isinstance(v, datetime):
        anchor = datetime(1899, 12, 30, tzinfo=timezone.utc)
        delta = v.replace(tzinfo=timezone.utc) - anchor
        return int(delta.days)
    return _to_int(v)


# Column indices (0-based) in each raffle type
STD_COLS = dict(num=1, name=2, date=3, txid=4, purchase=5,
                paid=6, free=7, total=8, hr=9, start=10, end=12)
HR_COLS  = dict(num=1, name=2, date=3, txid=4, purchase=5,
                paid=6, free=7, total=8, start=9, end=11)

STD_PRIZE_COLS = dict(active_at=16, prize=17, ticket=18, winner=19)
STD_PRIZE_ROW_RANGE = (17, 39)
STD_MINI_PRIZE_COLS = dict(prize=16, ticket=18, winner=19)
STD_MINI_ROW_RANGE = (59, 68)
HR_PRIZE_COLS = dict(active_at=15, prize=16, ticket=17, winner=18)
HR_PRIZE_ROW_RANGE = (14, 28)


def _parse_entry_std(row, raffle_id, conn):
    cols = STD_COLS
    if len(row) < 13:
        return None
    name = _norm_account(row[cols["name"]])
    if not name:
        return None
    occurred = _parse_dt(row[cols["date"]])
    purchase_raw = row[cols["purchase"]]
    paid = _to_int(row[cols["paid"]]) or 0
    free = _to_int(row[cols["free"]]) or 0
    hr = _to_int(row[cols["hr"]]) or 0
    start = _to_int(row[cols["start"]])
    end = _to_int(row[cols["end"]])
    txid_raw = row[cols["txid"]]
    purchase_int = _to_int(purchase_raw)
    purchase_str = str(purchase_raw).strip().upper() if purchase_raw is not None else ""

    if purchase_int is not None and purchase_int >= 1000:
        source = "bank_deposit"
        descriptor = None
        gold_amount = purchase_int
        txid = str(_to_int(txid_raw)) if txid_raw is not None else None
    elif purchase_str == "DONATION":
        source = "mail_donation"
        descriptor = "DONATION"
        gold_amount = _parse_donation_value_from_date_column(row[cols["date"]])
        if not occurred or occurred.year > 2100:
            occurred = None
        txid = None
    else:
        source = "event"
        descriptor = purchase_str if purchase_str else None
        gold_amount = None
        txid = None

    if occurred is None:
        return None

    user_id = upsert_user(conn, name)
    return RaffleEntry(
        raffle_id=raffle_id, user_id=user_id, source=source,
        occurred_at=occurred, gold_amount=gold_amount,
        paid_tickets=paid, free_tickets=free, high_roller_tickets=hr,
        descriptor=descriptor, source_transaction_id=txid,
        start_number=start, end_number=end, is_backfilled=True,
    )


def _parse_entry_hr(row, raffle_id, conn):
    cols = HR_COLS
    if len(row) < 12:
        return None
    name = _norm_account(row[cols["name"]])
    if not name:
        return None
    occurred = _parse_dt(row[cols["date"]])
    if occurred is None:
        return None
    purchase_int = _to_int(row[cols["purchase"]])
    paid = _to_int(row[cols["paid"]]) or 0
    free = _to_int(row[cols["free"]]) or 0
    start = _to_int(row[cols["start"]])
    end = _to_int(row[cols["end"]])
    txid_raw = row[cols["txid"]]
    txid = str(_to_int(txid_raw)) if txid_raw is not None else None

    user_id = upsert_user(conn, name)
    return RaffleEntry(
        raffle_id=raffle_id, user_id=user_id, source="high_roller_qualifier",
        occurred_at=occurred, gold_amount=purchase_int,
        paid_tickets=paid, free_tickets=free, high_roller_tickets=paid,
        descriptor=None, source_transaction_id=txid,
        start_number=start, end_number=end, is_backfilled=True,
    )


def _classify_prize(prize_value):
    if prize_value is None:
        return ("gold", None, None)
    g = _to_int(prize_value)
    if g is not None:
        return ("gold", g, None)
    return ("item", None, str(prize_value).strip())


def _parse_main_prizes_rows(all_rows, raffle_id, conn, cols, row_range):
    prize_count = winner_count = 0
    pos = 0
    for r in range(row_range[0], row_range[1] + 1):
        if r - 1 >= len(all_rows):
            break
        row = all_rows[r - 1]
        if not row:
            continue
        active_at_raw = row[cols["active_at"]] if len(row) > cols["active_at"] else None
        prize_raw = row[cols["prize"]] if len(row) > cols["prize"] else None
        ticket_raw = row[cols["ticket"]] if len(row) > cols["ticket"] else None
        winner_raw = row[cols["winner"]] if len(row) > cols["winner"] else None
        active_at = _to_int(active_at_raw)
        if active_at is None and prize_raw is None:
            continue
        prize_type, gold, item = _classify_prize(prize_raw)
        pos += 1
        prize_id = upsert_prize(conn, Prize(
            raffle_id=raffle_id, category="main", display_order=pos,
            active_at_ticket_count=active_at,
            prize_type=prize_type, gold_amount=gold, item_description=item,
        ))
        prize_count += 1
        ticket = _to_int(ticket_raw)
        winner = _norm_account(winner_raw)
        if ticket is not None and winner:
            user_id = upsert_user(conn, winner)
            upsert_winner(conn, prize_id=prize_id, user_id=user_id,
                          winning_ticket_number=ticket, source="backfill")
            winner_count += 1
    return prize_count, winner_count


def _parse_mini_prizes_std_rows(all_rows, raffle_id, conn):
    cols = STD_MINI_PRIZE_COLS
    prize_count = winner_count = 0
    pos = 0
    for r in range(STD_MINI_ROW_RANGE[0], STD_MINI_ROW_RANGE[1] + 1):
        if r - 1 >= len(all_rows):
            break
        row = all_rows[r - 1]
        if not row:
            continue
        prize_raw = row[cols["prize"]] if len(row) > cols["prize"] else None
        ticket_raw = row[cols["ticket"]] if len(row) > cols["ticket"] else None
        winner_raw = row[cols["winner"]] if len(row) > cols["winner"] else None
        if prize_raw is None and ticket_raw is None:
            continue
        prize_type, gold, item = _classify_prize(prize_raw)
        pos += 1
        prize_id = upsert_prize(conn, Prize(
            raffle_id=raffle_id, category="mini", display_order=pos,
            active_at_ticket_count=None,
            prize_type=prize_type, gold_amount=gold, item_description=item,
        ))
        prize_count += 1
        ticket = _to_int(ticket_raw)
        winner = _norm_account(winner_raw)
        if ticket is not None and winner:
            user_id = upsert_user(conn, winner)
            upsert_winner(conn, prize_id=prize_id, user_id=user_id,
                          winning_ticket_number=ticket, source="backfill")
            winner_count += 1
    return prize_count, winner_count


def backfill_one_tab(conn, ws, raffle_type, tab_name, counts, errors):
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return
    row2 = all_rows[1] if len(all_rows) >= 2 else None
    drawing_dt = _parse_dt(row2[10] if row2 and len(row2) > 10 else None)
    if drawing_dt is None:
        errors.append(f"  {raffle_type} {tab_name}: missing K2 drawing date")
        counts["rows_skipped"] += 1
        return

    raffle_id = upsert_raffle(conn, raffle_type=raffle_type,
                              drawing_date=drawing_dt.date(),
                              status="drawn", is_backfilled=True)

    parse_entry = _parse_entry_std if raffle_type == "standard" else _parse_entry_hr
    entries_added = 0
    for r_idx in range(3, min(203, len(all_rows))):
        row = all_rows[r_idx]
        if not row:
            continue
        try:
            e = parse_entry(row, raffle_id, conn)
        except Exception as exc:
            errors.append(f"  {raffle_type} {tab_name} R{r_idx+1}: entry parse error: {exc}")
            continue
        if e is None:
            continue
        try:
            insert_raffle_entry(conn, e)
            entries_added += 1
        except Exception as exc:
            errors.append(f"  {raffle_type} {tab_name} R{r_idx+1}: entry insert error: {exc}")

    if raffle_type == "standard":
        main_p, main_w = _parse_main_prizes_rows(all_rows, raffle_id, conn,
                                                 STD_PRIZE_COLS, STD_PRIZE_ROW_RANGE)
        mini_p, mini_w = _parse_mini_prizes_std_rows(all_rows, raffle_id, conn)
        prize_count = main_p + mini_p
        winner_count = main_w + mini_w
    else:
        main_p, main_w = _parse_main_prizes_rows(all_rows, raffle_id, conn,
                                                 HR_PRIZE_COLS, HR_PRIZE_ROW_RANGE)
        prize_count = main_p
        winner_count = main_w

    recompute_raffle_totals(conn, raffle_id)
    counts["rows_inserted"] += entries_added + prize_count + winner_count


def backfill_workbook(conn, path, raffle_type, counts, errors):
    print(f"\n=== {raffle_type} from {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    skipped = []
    n_tabs = 0
    for tab_name in wb.sheetnames:
        if not DATE_TAB_RE.fullmatch(tab_name):
            skipped.append(tab_name)
            continue
        ws = wb[tab_name]
        try:
            backfill_one_tab(conn, ws, raffle_type, tab_name, counts, errors)
            n_tabs += 1
        except Exception as exc:
            errors.append(f"  {raffle_type} {tab_name}: tab-level error: {exc}")
            counts["rows_skipped"] += 1
    print(f"  processed {n_tabs} tabs; skipped non-date: {skipped}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", required=True)
    ap.add_argument("--standard", required=True)
    ap.add_argument("--highroller", required=True)
    ap.add_argument("--schema", default=None)
    args = ap.parse_args()

    conn = open_db(args.db)
    if args.schema:
        apply_schema(conn, args.schema)

    errors = []
    with ingest_run(conn, "backfill_raffle", workbook_filename=str(args.standard)) as counts:
        conn.execute("BEGIN")
        try:
            backfill_workbook(conn, args.standard, "standard", counts, errors)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise

    with ingest_run(conn, "backfill_raffle", workbook_filename=str(args.highroller)) as counts:
        conn.execute("BEGIN")
        try:
            backfill_workbook(conn, args.highroller, "high_roller", counts, errors)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise

    if errors:
        print(f"\n=== {len(errors)} warnings ===")
        for e in errors[:25]:
            print(e)
        if len(errors) > 25:
            print(f"  ...and {len(errors) - 25} more")
    print("\nDone.")


if __name__ == "__main__":
    main()
