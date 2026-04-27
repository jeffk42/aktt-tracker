"""backfill.py - One-time historical backfill of guildstats.db from the
two archive Google-Sheet exports:

    AKTT Weekly Raw Data.xlsx     -> user_week_stats (per-user-per-week aggregates)
    AKTT Standard Raffle.xlsx     -> bank_transactions (raffle ticket purchases only)

After backfill, every historical trade week has a complete set of user_week_stats
rows (with is_backfilled=1), and every raffle ticket purchase since Dec 2022 has
a row in bank_transactions (with is_backfilled=1, type='dep_gold').

Non-raffle bank transactions (regular gold deposits, item donations, withdrawals)
are NOT recoverable from these spreadsheets - the spreadsheet only retains the
weekly per-user totals, which we store directly on user_week_stats.

Usage:
    python backfill.py --db guildstats.db \\
                       --donations "AKTT Weekly Raw Data.xlsx" \\
                       --raffle "AKTT Standard Raffle.xlsx" \\
                       --schema schema.sql
"""
from __future__ import annotations
import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from guildstats import (
    open_db, apply_schema,
    upsert_user, upsert_week, upsert_user_week_stats, upsert_bank_transaction,
    ingest_run, trade_week_ending, WeekStats, BankTxn,
    RAFFLE_DEPOSIT_MODIFIER,
)

DATE_TAB_RE = re.compile(r"^\d{6}$")


# --- helpers ----------------------------------------------------------------

def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _normalize_account(s) -> str | None:
    if s is None:
        return None
    s = str(s).strip()
    if not s.startswith("@"):
        return None
    return s


def _parse_excel_dt(v) -> datetime | None:
    """openpyxl gives us either datetime objects or strings depending on cell format."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    try:
        # Common spreadsheet string format
        return datetime.strptime(str(v), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


# --- donations backfill (per-user weekly aggregates) ------------------------

def backfill_donations(conn, workbook_path: str | Path, counts: dict) -> None:
    """Walk every MMDDYY tab in the donations workbook and UPSERT
    user_week_stats rows."""
    print(f"\n=== Backfilling donations from {workbook_path} ===")
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    skipped_tabs = []
    for tab_name in wb.sheetnames:
        if not DATE_TAB_RE.fullmatch(tab_name):
            skipped_tabs.append(tab_name)
            continue

        ws = wb[tab_name]
        # Row 2 col A holds the authoritative week-ending timestamp.
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        ts = _parse_excel_dt(row2[0])
        if ts is None:
            print(f"  warn: tab {tab_name} has no week-ending timestamp in A2; skipping",
                  file=sys.stderr)
            counts["rows_skipped"] += 1
            continue

        # Snap to the canonical Tuesday 19:00 UTC for consistency with live ingest.
        week_end = trade_week_ending(ts)
        week_id = upsert_week(conn, week_end)

        rows_for_tab = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            account = _normalize_account(row[0])
            if not account:
                continue
            rank      = _to_int(row[1])
            sales     = _to_int(row[2]) or 0
            taxes     = _to_int(row[3]) or 0
            deposits  = _to_int(row[4]) or 0
            raffle    = _to_int(row[5]) or 0
            donations = _to_int(row[6]) or 0
            purchases = _to_int(row[7]) or 0

            user_id = upsert_user(conn, account, excluded=False)
            upsert_user_week_stats(conn, WeekStats(
                user_id=user_id, week_id=week_id, rank=rank,
                sales=sales, taxes=taxes, purchases=purchases,
                total_deposits=deposits, total_raffle=raffle,
                total_donations=donations,
                is_backfilled=True,
            ))
            rows_for_tab += 1
            counts["rows_inserted"] += 1
        print(f"  tab {tab_name} ({week_end.date().isoformat()}): {rows_for_tab} rows")

    if skipped_tabs:
        print(f"  skipped non-date tabs: {skipped_tabs}")


# --- raffle backfill (per-transaction) --------------------------------------

def backfill_raffle(conn, workbook_path: str | Path, counts: dict) -> None:
    """Walk every MMDDYY tab in the raffle workbook and INSERT bank_transactions.
    Each tab represents one raffle drawing (Friday->Friday in ET); transactions
    inside may span 1 or 2 trade weeks depending on individual timestamps."""
    print(f"\n=== Backfilling raffle from {workbook_path} ===")
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    skipped_tabs = []
    for tab_name in wb.sheetnames:
        if not DATE_TAB_RE.fullmatch(tab_name):
            skipped_tabs.append(tab_name)
            continue

        ws = wb[tab_name]
        rows_for_tab = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Column layout (1-indexed):  A blank, B #, C name, D date, E txid,
            #                             F purchase amount, G-K computed
            if len(row) < 6:
                continue
            account = _normalize_account(row[2])
            if not account:
                continue
            occurred = _parse_excel_dt(row[3])
            if occurred is None:
                continue
            txid_raw = row[4]
            if txid_raw is None:
                continue
            txid = str(_to_int(txid_raw))  # strip any ".0" decimal
            purchase_amt = _to_int(row[5])
            if purchase_amt is None or purchase_amt <= 0:
                continue

            # Reconstruct the actual deposit amount including the +1 modifier
            # so the row is shape-compatible with live GBL ingest.
            gold_amount = purchase_amt + RAFFLE_DEPOSIT_MODIFIER

            week_end = trade_week_ending(occurred)
            week_id = upsert_week(conn, week_end)
            user_id = upsert_user(conn, account, excluded=False)
            result = upsert_bank_transaction(conn, BankTxn(
                transaction_id=txid,
                user_id=user_id,
                week_id=week_id,
                transaction_type="dep_gold",
                gold_amount=gold_amount,
                item_count=None,
                item_description=None,
                item_link=None,
                item_value=None,
                occurred_at=occurred,
                is_backfilled=True,
            ))
            if result == "inserted":
                counts["rows_inserted"] += 1
            else:
                counts["rows_skipped"] += 1
            rows_for_tab += 1
        print(f"  tab {tab_name}: {rows_for_tab} raffle rows")

    if skipped_tabs:
        print(f"  skipped non-date tabs: {skipped_tabs}")


# --- main -------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", required=True)
    ap.add_argument("--donations", required=True, help="AKTT Weekly Raw Data.xlsx")
    ap.add_argument("--raffle",    required=True, help="AKTT Standard Raffle.xlsx")
    ap.add_argument("--schema", default=None,
                    help="Optional path to schema.sql; applied before backfill if given")
    args = ap.parse_args()

    conn = open_db(args.db)
    if args.schema:
        apply_schema(conn, args.schema)

    # Donations first so user_week_stats is canonical; raffle backfill only
    # touches bank_transactions and won't disturb the aggregate totals.
    with ingest_run(conn, "backfill_donations",
                    workbook_filename=str(args.donations)) as counts:
        conn.execute("BEGIN")
        try:
            backfill_donations(conn, args.donations, counts)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise

    with ingest_run(conn, "backfill_raffle",
                    workbook_filename=str(args.raffle)) as counts:
        conn.execute("BEGIN")
        try:
            backfill_raffle(conn, args.raffle, counts)
            conn.execute("COMMIT")
        except Exception:
            conn.execute("ROLLBACK")
            raise

    print("\nDone.")


if __name__ == "__main__":
    main()
